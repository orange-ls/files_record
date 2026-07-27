#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
生成 Web 功能测试报告 (xlsx)
从 JSON 测试数据生成带截图的 Excel 测试报告

用法:
    python generate_report.py --data test_data.json --output 测试报告.xlsx
"""
import argparse
import json
import os
import sys
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as XlImage
    from openpyxl.utils import get_column_letter
except ImportError:
    print("错误: 需要安装 openpyxl 库")
    print("请执行: pip install openpyxl")
    sys.exit(1)

try:
    from PIL import Image as PILImage
except ImportError:
    PILImage = None


# 样式定义
HEADER_FONT = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill('solid', fgColor='4472C4')
PASS_FILL = PatternFill('solid', fgColor='C6EFCE')
PASS_FONT = Font(name='微软雅黑', size=10, color='006100')
FAIL_FILL = PatternFill('solid', fgColor='FFC7CE')
FAIL_FONT = Font(name='微软雅黑', size=10, color='9C0006')
NORMAL_FONT = Font(name='微软雅黑', size=10)
TITLE_FONT = Font(name='微软雅黑', bold=True, size=14)
SUBTITLE_FONT = Font(name='微软雅黑', size=10, color='666666')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_WRAP = Alignment(horizontal='left', vertical='center', wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 列配置: (标题, 宽度)
COLUMNS = [
    ('序号', 8),
    ('功能模块', 18),
    ('测试项', 20),
    ('测试描述', 35),
    ('预期效果', 35),
    ('实际效果', 35),
    ('系统截图', 40),
    ('测试结果', 12),
    ('问题描述', 35),
]


# 截图目标高度（像素），用于计算行高
SCREENSHOT_HEIGHT_PX = 200
SCREENSHOT_MAX_WIDTH_PX = 350


def resize_image_for_excel(img_path, max_width=SCREENSHOT_MAX_WIDTH_PX, max_height=SCREENSHOT_HEIGHT_PX):
    """调整图片大小以适合 Excel 单元格，返回 (宽, 高) 像素"""
    if PILImage is None:
        # 没有 PIL，使用默认尺寸
        return max_width, max_height
    try:
        with PILImage.open(img_path) as img:
            w, h = img.size
            ratio = min(max_width / w, max_height / h)
            return int(w * ratio), int(h * ratio)
    except Exception:
        return max_width, max_height


def generate_report(test_data, output_path):
    """生成测试报告 xlsx"""
    wb = Workbook()
    ws = wb.active
    ws.title = '测试报告'

    # 获取模块名和统计信息
    module_name = test_data[0].get('module_name', '未知模块') if test_data else '未知模块'
    total = len(test_data)
    passed = sum(1 for t in test_data if t.get('passed', False))
    failed = total - passed

    # 标题行
    ws.merge_cells('A1:I1')
    title_cell = ws['A1']
    title_cell.value = f'{module_name} - 功能测试报告'
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    # 摘要行
    ws.merge_cells('A2:I2')
    summary_cell = ws['A2']
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M')
    summary_cell.value = f'测试时间: {now_str}  |  总计: {total} 项  |  通过: {passed} 项  |  不通过: {failed} 项'
    summary_cell.font = SUBTITLE_FONT
    summary_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 25

    # 空行
    ws.row_dimensions[3].height = 10

    # 表头 (第4行)
    header_row = 4
    for col_idx, (title, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[header_row].height = 30

    # 数据行
    for i, tc in enumerate(test_data):
        row = header_row + 1 + i
        test_passed = tc.get('passed', False)

        values = [
            tc.get('test_id', i + 1),
            tc.get('module_name', ''),
            tc.get('test_item', ''),
            tc.get('test_description', ''),
            tc.get('expected_result', ''),
            tc.get('actual_result', ''),
            '',  # 截图列，后面插入图片
            '通过' if test_passed else '不通过',
            tc.get('issue', ''),
        ]

        row_height = 25  # 默认行高
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = NORMAL_FONT
            cell.alignment = LEFT_WRAP if col_idx in (4, 5, 6, 9) else CENTER
            cell.border = THIN_BORDER

        # 测试结果列样式
        result_cell = ws.cell(row=row, column=8)
        if test_passed:
            result_cell.fill = PASS_FILL
            result_cell.font = PASS_FONT
        else:
            result_cell.fill = FAIL_FILL
            result_cell.font = FAIL_FONT

        # 插入截图
        screenshot_path = tc.get('screenshot_path', '')
        if screenshot_path and os.path.isfile(screenshot_path):
            try:
                img_w, img_h = resize_image_for_excel(screenshot_path)
                img = XlImage(screenshot_path)
                img.width = img_w
                img.height = img_h
                # 图片锚定到 G 列 (第7列)
                cell_ref = f'G{row}'
                ws.add_image(img, cell_ref)
                # 调整行高以适应图片
                row_height = max(row_height, img_h * 0.75 + 10)
            except Exception as e:
                ws.cell(row=row, column=7, value=f'截图加载失败: {str(e)}')

        ws.row_dimensions[row].height = row_height

    # 冻结表头
    ws.freeze_panes = 'A5'

    # 保存
    wb.save(output_path)
    return {
        'status': 'success',
        'output': output_path,
        'total': total,
        'passed': passed,
        'failed': failed,
    }


def main():
    parser = argparse.ArgumentParser(description='生成 Web 功能测试报告 (xlsx)')
    parser.add_argument('--data', required=True, help='测试数据 JSON 文件路径')
    parser.add_argument('--output', required=True, help='输出 xlsx 文件路径')
    args = parser.parse_args()

    with open(args.data, 'r', encoding='utf-8') as f:
        test_data = json.load(f)

    if not test_data:
        print('错误: 测试数据为空')
        sys.exit(1)

    result = generate_report(test_data, args.output)
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == '__main__':
    main()

"""
通用 Excel 导出工具
统一样式：黄色表头 + 微软雅黑，数据行白底 + 宋体。
支持两种数据源：
  1. fetcher query() 返回的 records（list[dict]）
  2. MCP SQL 查询返回的 records（list[dict]）

用法：
    from ._export import export_to_excel
    path = export_to_excel(
        records=result['records'],
        fields=['material_code', 'material_desc', ...],
        field_labels={'material_code': '物料代码', ...},
        sheet_title='BOM总表',
    )
"""
import datetime
import logging
import os
import re
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter

_logger = logging.getLogger(__name__)

# 桌面路径（跨平台兼容）
_DESKTOP = os.path.join(os.path.expanduser('~'), 'Desktop')

# ── 统一样式定义 ──
_THIN = Side(border_style='thin')
_BORDER = Border(top=_THIN, left=_THIN, right=_THIN, bottom=_THIN)
_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)

# 表头样式：黄色背景 + 微软雅黑加粗
_HEAD_FONT = Font(name='微软雅黑', bold=True, color='000000', size=12)
_HEAD_FILL = PatternFill(fill_type='solid', fgColor='FFFF00')

# 数据行样式：白色背景 + 宋体
_DATA_FONT = Font(name='宋体', size=11, color='000000')
_DATA_FILL = PatternFill(fill_type='solid', fgColor='FFFFFF')


def _safe_value(value):
    """
    处理单元格值：
    - None → 空字符串
    - 字符串以 '=' 开头 → 前加空格（防止 Excel 公式注入）
    - date/datetime → 格式化为字符串
    - 其他 → 原样返回
    """
    if value is None:
        return ''
    if isinstance(value, str):
        return re.sub(r'^=', ' =', value) if value else ''
    if isinstance(value, datetime.datetime):
        return value.strftime('%Y-%m-%d %H:%M:%S')
    if isinstance(value, datetime.date):
        return value.strftime('%Y-%m-%d')
    return value


def export_to_excel(
    records: list,
    fields: list,
    field_labels: dict,
    sheet_title: str,
    file_name: str = None,
    save_dir: str = None,
) -> str:
    """
    将数据导出为 Excel 文件。

    :param records: 数据列表，每条为 dict
    :param fields: 导出字段列表（按顺序），对应 records 中的 key
    :param field_labels: 字段名 → 中文表头映射
    :param sheet_title: Sheet 页名称（也用于默认文件名）
    :param file_name: 自定义文件名（不含路径），默认为 "{sheet_title}_{日期}.xlsx"
    :param save_dir: 保存目录，默认为桌面
    :return: 保存的文件完整路径
    """
    save_dir = save_dir or _DESKTOP
    os.makedirs(save_dir, exist_ok=True)

    if not file_name:
        timestamp = datetime.datetime.now().strftime('%Y%m%d')
        file_name = f'{sheet_title}_{timestamp}.xlsx'

    file_path = os.path.join(save_dir, file_name)

    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    # 注册命名样式（避免重复创建）
    head_style = NamedStyle(
        name='head_style',
        font=_HEAD_FONT, fill=_HEAD_FILL,
        border=_BORDER, alignment=_ALIGNMENT,
    )
    data_style = NamedStyle(
        name='data_style',
        font=_DATA_FONT, fill=_DATA_FILL,
        border=_BORDER, alignment=_ALIGNMENT,
    )
    wb.add_named_style(head_style)
    wb.add_named_style(data_style)

    # 写入表头
    headers = [field_labels.get(f, f) for f in fields]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(1, col_idx, value=header).style = 'head_style'

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 写入数据行
    for row_idx, record in enumerate(records, 2):
        for col_idx, field in enumerate(fields, 1):
            value = _safe_value(record.get(field))
            ws.cell(row_idx, col_idx, value=value).style = 'data_style'

    # 自动列宽（基于表头长度，最小 13，最大 50）
    for col_idx, header in enumerate(headers, 1):
        # 中文字符按 2 宽度计算
        char_width = sum(2 if ord(c) > 127 else 1 for c in header)
        width = max(13, min(50, char_width + 4))
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 保存文件
    wb.save(file_path)
    _logger.info('[xc_spare_parts] Excel 导出完成：%s，共 %d 条', file_path, len(records))

    return file_path

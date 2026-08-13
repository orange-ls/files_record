# -*- coding: utf-8 -*-
"""
备货-销售差异数据导出脚本

用法:
    python export_stocking_sales_variance.py

功能:
    1. 导出search_read()返回的主表数据为Excel，文件名: 备货-销售差异_时间戳.xlsx
    2. 对每条记录调用get_all_variance_table()查询数据，导出4个子表的Excel
"""

import io
import os
import sys
import logging
import psycopg2
from psycopg2.extras import RealDictCursor
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Side, Border, PatternFill, Font, NamedStyle
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
_logger = logging.getLogger(__name__)

# ==================== 数据库配置 ====================
DB_CONFIG = {
    'host': '10.0.23.199',
    'port': 5432,
    'user': 'xc',
    'password': 'Dcxc7888$',
    'database': 'xc_materiel',
}

# 导出文件存放目录
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'output')


def get_db_connection():
    """获取数据库连接"""
    return psycopg2.connect(**DB_CONFIG)


# ==================== SQL查询（与 stocking_sales_variance.py 中一致） ====================

SEARCH_READ_SQL = '''
    WITH quotation_data AS (
        SELECT DISTINCT x.crm_no, x.crm_name as project_name
        FROM xc_quotation x
        INNER JOIN vm_dcn_xsmx v ON x.crm_no = v.crmxmh AND x.crm_name = v.lyjh
        WHERE v.crmxmh IS NOT NULL AND v.crmxmh <> ''
            AND v.lyjh IS NOT NULL AND v.lyjh <> ''
            AND x.crm_no IS NOT NULL AND x.crm_no <> ''
            AND x.crm_name IS NOT NULL AND x.crm_name <> ''
    )
    SELECT
        qd.crm_no, qd.project_name,
        cnoo.project_industry as industry_one,
        cnoo.sales_name, cnoo.create_time
    FROM quotation_data qd
    LEFT JOIN crm_new_opportunity_obj cnoo ON qd.crm_no = cnoo.crm_project_num
    WHERE cnoo.create_time > '2025-01-01'
    ORDER BY cnoo.create_time DESC
'''


def search_read(cr):
    """查询主表数据（与 model 中 search_read 逻辑一致）"""
    cr.execute(SEARCH_READ_SQL)
    return cr.fetchall()


def get_mat_type(mat_name):
    """根据物料名称判断物料类型"""
    if not mat_name:
        return None
    if '固件' in mat_name or '包装' in mat_name:
        return '辅助料信息'
    elif '定制' in mat_name:
        return '定制化信息'
    elif mat_name.endswith('服务'):
        return '服务类信息'
    return '基础物料'


def transform_sap_no(sap_no):
    """将物料号规范为 302-123456 的格式"""
    if not sap_no:
        return ''
    new_no = str(sap_no).lstrip('0')
    if len(new_no) > 6:
        return f'{new_no[:-6]}-{new_no[-6:]}'
    return new_no


def get_all_variance_table(cr, crm_no, project_name):
    """
    一次性获取Form视图中所有4个表的数据（与 model 中逻辑一致）
    """
    # 获取报价单号
    cr.execute(
        'SELECT quot_no FROM xc_quotation WHERE crm_no = %s AND crm_name = %s',
        (crm_no, project_name)
    )
    result = cr.fetchone()
    if not result:
        return {}
    quot_no = result['quot_no']

    # ========== 1. 获取备货配置BOM ==========
    total_inf_sql = '''
        SELECT id,
               ROW_NUMBER() OVER (ORDER BY "sequence" ASC, id ASC) as num,
               pro_sap_no, pro_type, pro_numb
        FROM xc_product
        WHERE quot_no = %s AND del_pro_flag = '0'
    '''
    cr.execute(total_inf_sql, (quot_no,))
    total_inf = cr.fetchall()

    total_inf_map = {
        ti['id']: {
            'material_type': ti['num'],
            'product_name': ti['pro_sap_no'],
            'material_code': ti['pro_type'],
            'quantity_sum': ti['pro_numb']
        }
        for ti in total_inf
    }

    config_bom_list = []
    config_bom_material = {}

    # 物料类型映射SQL
    purchase_type_sql = '''
        SELECT sap_no,
               CASE MIN(purchase_type)
                   WHEN '0' THEN '华为采'
                   ELSE '外采'
               END as purchase_type
        FROM "xc_price_actuarial"
        WHERE quot_no = %s
        GROUP BY sap_no
    '''
    cr.execute(purchase_type_sql, (quot_no,))
    purchase_type_map = {row['sap_no']: row['purchase_type'] for row in cr.fetchall()}

    for t_id, ti in total_inf_map.items():
        config_bom_list.append(ti)

        config_detail_sql = '''
            SELECT
                xcpd.mat_name as material_type,
                xcpdt.transfer_mat_name as product_name,
                xcpdt.transfer_sap_no as material_code,
                xcpdt.transfer_spec as spec,
                xcpd.quantity as quantity_per,
                xcpdt.specify_purchase,
                xcpd.sale_comment
            FROM xc_product_detail xcpd
            LEFT JOIN xc_product_detail_transfer xcpdt ON xcpdt.product_detail_id = xcpd.id
            WHERE xcpd.pro_id = %s
                AND xcpd.sap_no IS NOT NULL
                AND xcpd.sap_no != ''
                AND xcpd.quantity > 0
                AND xcpd.mat_type != '主机型号'
            ORDER BY xcpd.number_sort, xcpd.id
        '''
        cr.execute(config_detail_sql, (t_id,))
        config_records = cr.fetchall()

        for cre in config_records:
            cre['material_type'] = get_mat_type(cre.get('material_type', ''))
            cre['quantity_sum'] = cre['quantity_per'] * ti['quantity_sum']
            cre['purchase_type'] = purchase_type_map.get(cre.get('material_code'), '')
            config_bom_list.append(cre)

            # 1.2、备货配置BOM 物料汇总
            code = cre['material_code']
            if code:
                if code not in config_bom_material:
                    config_bom_material[code] = {'spec': cre.get('spec', ''), 'quantity_sum': 0}
                config_bom_material[code]['quantity_sum'] += cre['quantity_sum']

    # ========== 2. 获取销售订单明细 ==========
    sales_order_sql = '''
        SELECT xsddm as sale_no, xsdhxm as sale_order_line_item,
               jhddm as delivery_no, jhdhxm as delivery_note_line_item,
               wldm as material_code, wlmc as material_name, sl as quantity
        FROM "vm_dcn_xsmx"
        WHERE crmxmh = %s AND lyjh = %s
        ORDER BY xsddm, wldm ASC
    '''
    cr.execute(sales_order_sql, (crm_no, project_name))
    sales_order_list = cr.fetchall()

    # 转换物料代码格式
    # 2.2 销售订单明细 物料汇总
    sales_order_material = {}
    for sol in sales_order_list:
        sol['material_code'] = transform_sap_no(sol['material_code'])
        code = sol['material_code']
        qty = int(sol.get('quantity', 0))
        sales_order_material[code] = sales_order_material.get(code, 0) + qty

    # ========== 2.3、获取产品BOM ==========
    product_bom_map = []
    sales_material_sum = {}
    i = 1
    for code, qty in sales_order_material.items():
        # 查询该物料的BOM组件
        product_bom_sql = '''
            SELECT material_code, material_name, factory_code, project_type,
                   bom_pro_num, bom_component_code, bom_component_name, component_num
            FROM material_find_manage
            WHERE material_code = %s
            ORDER BY bom_component_code ASC
        '''
        cr.execute(product_bom_sql, (code,))
        product_bom_list = cr.fetchall()

        if product_bom_list:
            product_bom_map.append({'material_code': i, 'component_num': qty})
            i = i + 1
            for pbl in product_bom_list:
                pbl['bom_pro_num'] = str(pbl['bom_pro_num']).lstrip('0')
                pbl['quantity_sum'] = pbl['component_num'] * qty
                product_bom_map.append(pbl)

                # 2.4 销售物料汇总
                component_code = pbl['bom_component_code']
                if component_code:
                    if component_code not in sales_material_sum:
                        sales_material_sum[component_code] = {'spec': pbl.get('bom_component_name', ''), 'quantity_sum': 0}
                    sales_material_sum[component_code]['quantity_sum'] += pbl['quantity_sum']

    # ========== 3. 计算备货-销售差异汇总 ==========
    material_code_list = sorted(
        set(config_bom_material.keys()) | set(sales_material_sum.keys()),
        key=lambda x: (len(x), x)
    )

    variance_sum = [
        {
            'material_code': code,
            'material_name': config_bom_material.get(code, {}).get('spec') or
                             sales_material_sum.get(code, {}).get('spec', ''),
            'stocking_quantity': config_bom_material.get(code, {}).get('quantity_sum', 0),
            'sale_quantity': sales_material_sum.get(code, {}).get('quantity_sum', 0),
            'difference': (config_bom_material.get(code, {}).get('quantity_sum', 0) -
                          sales_material_sum.get(code, {}).get('quantity_sum', 0))
        }
        for code in material_code_list
    ]

    return {
        '备货配置BOM': config_bom_list,
        '销售订单明细': sales_order_list,
        '产品BOM': product_bom_map,
        '备货-销售差异汇总': variance_sum
    }


# ==================== Excel 导出 ====================

def create_head_style():
    """创建表头样式"""
    thin = Side(border_style="thin")
    font = Font(name="微软雅黑", size=11)
    fill = PatternFill(fill_type='solid', fgColor="C0C0C0")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    return NamedStyle(name="head_style", font=font, border=border, alignment=alignment, fill=fill)


def create_sheet(workbook, sheet_name, headers, data, field_names, head_style):
    """创建一个sheet并填充数据"""
    sheet = workbook.create_sheet(title=sheet_name)

    # 写入表头
    for col_idx, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.style = head_style
        sheet.column_dimensions[get_column_letter(col_idx)].width = 20

    # 写入数据
    for row_idx, record in enumerate(data, 2):
        for col_idx, field_name in enumerate(field_names, 1):
            value = record.get(field_name)
            if value is None:
                value = ''
            sheet.cell(row=row_idx, column=col_idx, value=value)


def export_main_table(records, output_dir):
    """
    导出主表数据（search_read结果）为Excel
    文件名: 备货-销售差异_时间戳.xlsx
    """
    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
    filename = f'备货-销售差异_{timestamp}.xlsx'
    filepath = os.path.join(output_dir, filename)

    workbook = Workbook()
    if 'Sheet' in workbook.sheetnames:
        del workbook['Sheet']

    head_style = create_head_style()

    create_sheet(
        workbook,
        '备货-销售差异',
        ['CRM立项编号', '项目名称', '项目行业一级', '销售员', '创建时间'],
        records,
        ['crm_no', 'project_name', 'industry_one', 'sales_name', 'create_time'],
        head_style
    )

    workbook.save(filepath)
    _logger.info(f"主表数据已导出: {filepath} (共{len(records)}条记录)")
    return filepath


def export_detail_table(cr, crm_no, project_name, output_dir):
    """
    导出单条记录的4个子表数据为Excel
    文件名: crm_no_项目名称.xlsx
    """
    all_data = get_all_variance_table(cr, crm_no, project_name)
    if not all_data:
        _logger.warning(f"未获取到数据: {crm_no} - {project_name}")
        return None

    filename = f"{crm_no}_{project_name}.xlsx"
    filepath = os.path.join(output_dir, filename)

    workbook = Workbook()
    if 'Sheet' in workbook.sheetnames:
        del workbook['Sheet']

    head_style = create_head_style()

    # 1. 备货-销售差异汇总
    create_sheet(
        workbook,
        '备货-销售差异汇总',
        ['物料代码', '物料名称', '备货总数量', '销售总数量', '差异(备货-销售)'],
        all_data.get('备货-销售差异汇总', []),
        ['material_code', 'material_name', 'stocking_quantity', 'sale_quantity', 'difference'],
        head_style
    )

    # 2. 备货配置BOM
    create_sheet(
        workbook,
        '备货配置BOM',
        ['物料类别', '产品名称', '料号', '描述', '单套数量', '总数量', '采购类型', '是否指定物料', '特殊说明'],
        all_data.get('备货配置BOM', []),
        ['material_type', 'product_name', 'material_code', 'spec', 'quantity_per', 'quantity_sum',
         'purchase_type', 'specify_purchase', 'sale_comment'],
        head_style
    )

    # 3. 销售订单明细
    create_sheet(
        workbook,
        '销售订单明细',
        ['销售订单号', '销售订单行项目', '交货单号', '交货单行项目', '物料代码', '物料名称', '数量'],
        all_data.get('销售订单明细', []),
        ['sale_no', 'sale_order_line_item', 'delivery_no', 'delivery_note_line_item',
         'material_code', 'material_name', 'quantity'],
        head_style
    )

    # 4. 产品BOM明细
    create_sheet(
        workbook,
        '产品BOM明细',
        ['物料代码', '物料名称', '工厂代码', '项目类别', 'BOM项目号', 'BOM组件代码', 'BOM组件名称', '组件数量', '总数量'],
        all_data.get('产品BOM', []),
        ['material_code', 'material_name', 'factory_code', 'project_type', 'bom_pro_num',
         'bom_component_code', 'bom_component_name', 'component_num', 'quantity_sum'],
        head_style
    )

    workbook.save(filepath)
    _logger.info(f"详情数据已导出: {filepath}")
    return filepath


def main():
    """主函数"""
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    _logger.info(f"导出目录: {OUTPUT_DIR}")

    conn = get_db_connection()
    try:
        cr = conn.cursor(cursor_factory=RealDictCursor)

        # 1. 查询主表数据
        _logger.info("正在查询主表数据...")
        records = search_read(cr)
        _logger.info(f"共查询到 {len(records)} 条记录")

        if not records:
            _logger.info("无数据可导出")
            return

        # 2. 导出主表Excel
        main_filepath = export_main_table(records, OUTPUT_DIR)

        # 3. 逐条导出详情Excel
        _logger.info("正在导出详情数据...")
        detail_dir = os.path.join(OUTPUT_DIR, 'details')
        os.makedirs(detail_dir, exist_ok=True)

        success_count = 0
        failed_records = []
        for idx, record in enumerate(records, 1):
            crm_no = record['crm_no']
            project_name = record['project_name']
            _logger.info(f"[{idx}/{len(records)}] 正在处理: {crm_no} - {project_name}")
            try:
                filepath = export_detail_table(cr, crm_no, project_name, detail_dir)
                if filepath:
                    success_count += 1
                else:
                    failed_records.append((crm_no, project_name, '无数据(quot_no未找到或数据为空)'))
            except Exception as e:
                failed_records.append((crm_no, project_name, str(e)))
                _logger.error(f"导出失败 [{crm_no} - {project_name}]: {e}")

        _logger.info(f"导出完成! 主表: {main_filepath}, 详情成功: {success_count}/{len(records)}")

        if failed_records:
            _logger.warning(f"共 {len(failed_records)} 条记录导出失败:")
            for crm_no, project_name, reason in failed_records:
                _logger.warning(f"  {crm_no} - {project_name}: {reason}")

            # 将失败记录导出到Excel
            fail_filename = f'导出失败记录_{datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx'
            fail_filepath = os.path.join(OUTPUT_DIR, fail_filename)
            wb = Workbook()
            ws = wb.active
            ws.title = '失败记录'
            ws.append(['CRM立项编号', '项目名称', '失败原因'])
            for crm_no, project_name, reason in failed_records:
                ws.append([crm_no, project_name, reason])
            wb.save(fail_filepath)
            _logger.info(f"失败记录已导出: {fail_filepath}")

    finally:
        conn.close()


if __name__ == '__main__':
    main()

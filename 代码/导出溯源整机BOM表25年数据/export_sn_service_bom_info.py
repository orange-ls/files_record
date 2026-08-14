# -*- coding: utf-8 -*-
# @Time: 2026-08-14
# @File: export_sn_service_bom_info.py
# @Desc: sn.service.bom.info(服务器-BOM组件信息) 高效导出脚本(独立运行, 不依赖 Odoo ORM)
#
# 适用场景: 按"产品交付时间(deliver_time)"区间批量导出 BOM 组件信息至 .xlsx, 数据量百万级。
#
# 用法示例(在项目根目录执行):
#   python xc_interface\export_data\export_sn_service_bom_info.py                        # 默认连内置生产库, 导出 2025-01-01 ~ 2025-12-31
#   python xc_interface\export_data\export_sn_service_bom_info.py --rows-per-file 500000 # 每卷最多 50 万行
#   python xc_interface\export_data\export_sn_service_bom_info.py --start 2025-07-01 --end 2025-12-31
#
# 性能设计要点:
#   1. keyset 分页(id > 上批末行 ORDER BY id LIMIT n): 每批独立短事务流式拉取,
#      客户端内存占用与数据总量无关, 且无长事务(不会被超时中断作废/不阻碍 VACUUM);
#   2. 过滤条件(WHERE)全部下推数据库层, 仅传输符合条件的数据, 日期参数化防 SQL 注入;
#   3. openpyxl write_only(只写)模式流式生成 xlsx, 内存恒定;
#   4. xlsx 单表上限 1048576 行, 超限自动分卷(默认每卷 100 万行数据 + 1 行表头);
#   5. 断点续传: 每次分卷完成即写 checkpoint, 中断后重跑脚本自动从断点继续
#      (已封卷文件保留, 未封卷部分重导); 单批连接中断自动重连重试;
#   6. 全程进度/速率/ETA 显示; 写入 .part 临时文件, 全部完成后原子重命名, 异常自动清理。

import argparse
import bisect
import json
import logging
import os
import sys
import time
from datetime import date, datetime

import psycopg2
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

LOG = logging.getLogger('export_bom')

# ----------------------------- 常量配置 -----------------------------
TABLE_NAME = 'sn_service.bom.info'.replace('.', '_')  # 模型 sn.service.bom.info 对应的数据表

# 导出字段: (数据库列名, Excel表头, 列宽); 如需调整导出列, 增删本列表即可
FIELDS = [
    ('unique_id',            '系统编号',           20),
    ('delivery_no',          '交货单号',           20),
    ('complete_sn',          '整机SN',             22),
    ('complete_unique_id',   '关联整机编号',       20),
    ('model',                '机型',               16),
    ('material_code',        '物料代码',           16),
    ('material_desc',        '物料描述',           32),
    ('bom_sn',               'BOM组件SN',          22),
    ('material_num',         '物料数量',           10),
    ('deliver_time',         '产品交付时间',       14),
    ('service_start_time',   '服务开始时间',       14),
    ('service_end_time',     '服务结束时间',       14),
    ('service_product_type', '服务产品类别',       14),
    ('customer_name',        '签约客户名称',       28),
    ('customer_level',       '客户等级',           10),
    ('proj_name',            '项目名称',           24),
    ('crm_no',               'CRM立项编号',        18),
    ('sale',                 '销售员',             12),
    ('sales',                '项目审批一线工程师', 18),
    ('province',             '所在省份',           10),
    ('city',                 '所在城市',           12),
    ('delivery_address',     '交付地址',           30),
    ('material_supplier',    '物料供应商',         16),
    ('purchase_time',        '采购时间',           14),
    ('is_add',               '是否项目增配',       12),
    ('add_proj_name',        '增配整机项目名',     20),
    ('add_crm_no',           '增配整机CRM立项编号', 22),
    ('sales_order_type',     '销售单类型',         12),
    ('remark',               '备注',               20),
]

BATCH_SIZE_DEFAULT = 10000       # 每批点查的行数(单批内存约几十MB以内)
ROWS_PER_FILE_DEFAULT = 1000000  # 每个xlsx文件最大"数据行"数(须小于Excel单表上限1048575)
EXCEL_SHEET_MAX_ROWS = 1048576   # xlsx 单个工作表最大行数(含表头, 硬性限制)

# 表头样式(仅表头设置样式; 正文不加样式, 以保障百万级行的写入性能与文件体积)
FONT_HEADER = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
FILL_HEADER = PatternFill('solid', fgColor='4472C4')
ALIGN_HEADER = Alignment(horizontal='center', vertical='center')

# 内置数据库连接配置(生产库 xc_materiel, 硬编码为唯一连接来源, 不接受任何外部覆盖)
# 注意: 内含敏感信息, 请勿将本文件提交至公开仓库或随意外发
DB_CONFIG = {
    'dbname': 'xc_materiel',
    'user': 'xc',
    'password': 'Dcxc7888$',
    'host': '10.0.23.199',
    'port': 5432,
}


class RollingExcelWriter:
    """分卷 xlsx 流式写入器(openpyxl 只写模式)

    - 单卷达到 rows_per_file 上限后自动保存当前文件并开启新卷;
    - 先写 .part 临时文件, 落盘完成后原子重命名(os.replace)为正式文件;
    - 仅表头应用样式与自动筛选, 正文保持默认样式以保障百万级行写入性能。
    """

    def __init__(self, output_dir, base_name, headers, widths, rows_per_file, start_file_idx=1):
        self.output_dir = output_dir
        self.base_name = base_name
        self.headers = headers
        self.widths = widths
        self.rows_per_file = rows_per_file
        self.files = []  # 已完成文件列表: [(最终路径, 数据行数)]
        self._wb = None
        self._ws = None
        self._tmp_path = None
        # 支持断点续传: 从指定卷号继续开卷(如上次已封卷至 part2, 则续传从 part3 开始)
        self._file_idx = start_file_idx - 1
        self._rows_in_file = 0

    def _open_volume(self):
        """开启新的一卷文件"""
        self._file_idx += 1
        self._wb = Workbook(write_only=True)  # 只写模式: 流式序列化, 内存占用恒定
        ws = self._wb.create_sheet(title='BOM组件信息')
        # 列宽/冻结窗格必须在写入首行数据之前设置(write_only 模式限制)
        for i, w in enumerate(self.widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = 'A2'
        # 表头行(带样式)
        header_cells = []
        for title in self.headers:
            cell = WriteOnlyCell(ws, value=title)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = ALIGN_HEADER
            header_cells.append(cell)
        ws.append(header_cells)
        self._ws = ws
        self._rows_in_file = 0
        self._tmp_path = os.path.join(
            self.output_dir, '%s_part%d.xlsx.part' % (self.base_name, self._file_idx))
        LOG.info('开启新分卷: %s (第 %d 卷)', self._tmp_path, self._file_idx)

    def _close_volume(self):
        """保存并封卷当前文件"""
        # 自动筛选区域 = 表头行 + 已写入数据行
        self._ws.auto_filter.ref = 'A1:%s%d' % (
            get_column_letter(len(self.headers)), self._rows_in_file + 1)
        self._wb.save(self._tmp_path)
        final_path = self._tmp_path[:-len('.part')]
        os.replace(self._tmp_path, final_path)  # 原子重命名, 避免暴露半成品文件
        self.files.append((final_path, self._rows_in_file))
        LOG.info('已生成文件: %s (数据 %d 行)', final_path, self._rows_in_file)
        self._wb = self._ws = self._tmp_path = None

    def write_rows(self, rows):
        """批量写入多行, 达到单卷上限自动轮转新文件"""
        for row in rows:
            if self._ws is None or self._rows_in_file >= self.rows_per_file:
                if self._ws is not None:
                    self._close_volume()
                self._open_volume()
            self._ws.append(row)
            self._rows_in_file += 1

    def close(self):
        """正常结束: 保存最后一个未封卷的文件"""
        if self._ws is not None:
            self._close_volume()

    def _release_volume_resources(self):
        """释放当前未封卷分卷占用的 openpyxl 内部资源(生成器/临时文件句柄)

        按 openpyxl write_only 内部结构依次收尾:
        关闭写行生成器 -> 关闭 XML 流生成器(释放临时文件句柄) ->
        删除 XML 临时文件并从全局清单注销 -> 关闭工作簿 zip 句柄;
        任一步骤的底层异常仅吞掉, 不影响中断主流程。
        """
        if self._ws is not None:
            try:
                rows = getattr(self._ws, '_rows', None)
                if rows is not None:
                    # 闭合 sheetData 元素, 生成器挂回 xf 等待收尾
                    rows.close()
            except BaseException:
                pass
            writer = getattr(self._ws, '_writer', None)
            if writer is not None:
                try:
                    # 关闭 get_stream 生成器, 释放其持有的 XML 临时文件句柄
                    writer.close()
                except BaseException:
                    pass
                try:
                    # 删除 XML 临时文件并从 openpyxl 全局清单注销,
                    # 避免 atexit 阶段 Windows 下文件仍被占用报 PermissionError
                    writer.cleanup()
                except (OSError, ValueError):
                    pass
        if self._wb is not None:
            try:
                # 关闭工作簿 zip 句柄(NamedTemporaryFile 随句柄关闭自动删除)
                self._wb.close()
            except BaseException:
                pass
        self._wb = self._ws = None

    def abort(self):
        """异常中断: 释放未完成分卷的内部资源并丢弃文件(已完成的分卷保留)"""
        self._release_volume_resources()
        if self._tmp_path and os.path.exists(self._tmp_path):
            try:
                os.remove(self._tmp_path)
                LOG.warning('已清理未完成的临时文件: %s', self._tmp_path)
            except OSError:
                LOG.exception('清理临时文件失败: %s', self._tmp_path)
            self._tmp_path = None


def parse_date(text):
    """命令行日期参数解析: YYYY-MM-DD"""
    try:
        return datetime.strptime(text, '%Y-%m-%d').date()
    except ValueError:
        raise argparse.ArgumentTypeError('日期格式错误: %r, 应为 YYYY-MM-DD' % text)


def parse_args():
    parser = argparse.ArgumentParser(
        description='导出 sn_service_bom_info 表数据至 xlsx(按产品交付时间区间筛选, 支持百万级数据量)')
    parser.add_argument('--start', type=parse_date, default=date(2025, 1, 1),
                        help='交付时间起始日期(默认 2025-01-01)')
    parser.add_argument('--end', type=parse_date, default=date(2025, 12, 31),
                        help='交付时间截止日期(默认 2025-12-31, 含当天)')
    parser.add_argument('--output', default=os.path.join(os.path.dirname(os.path.abspath(__file__)), 'output'),
                        help='输出目录(默认脚本目录下 output/)')
    parser.add_argument('--batch-size', type=int, default=BATCH_SIZE_DEFAULT,
                        help='每次从数据库拉取的行数(默认 %d)' % BATCH_SIZE_DEFAULT)
    parser.add_argument('--rows-per-file', type=int, default=ROWS_PER_FILE_DEFAULT,
                        help='每个xlsx文件最大数据行数(默认 %d)' % ROWS_PER_FILE_DEFAULT)
    parser.add_argument('--include-deleted', action='store_true',
                        help='包含逻辑删除数据(del_flag != 0), 默认仅导出有效数据')
    parser.add_argument('--no-count', action='store_true',
                        help='跳过前置 COUNT 总数查询(牺牲进度百分比, 换取启动速度)')
    parser.add_argument('--max-retries', type=int, default=3,
                        help='单批查询连接中断后的重连重试次数(默认 3, 退避 10/30/90 秒)')
    parser.add_argument('--restart', action='store_true',
                        help='忽略已有断点记录从头导出(旧分卷文件需手动清理)')
    args = parser.parse_args()

    # 参数合法性校验
    if args.start > args.end:
        parser.error('--start 不能晚于 --end')
    if args.batch_size < 1:
        parser.error('--batch-size 必须为正整数')
    if not 1 <= args.rows_per_file <= EXCEL_SHEET_MAX_ROWS - 1:
        parser.error('--rows-per-file 必须在 1 ~ %d 之间(Excel单表行数上限)' % (EXCEL_SHEET_MAX_ROWS - 1))
    return args


def build_where(start, end, include_deleted):
    """构建参数化的 WHERE 子句(列名为脚本内置常量, 日期值走 %s 占位符防注入)"""
    where = 'WHERE deliver_time >= %s AND deliver_time <= %s'
    params = [start, end]
    if not include_deleted:
        # 项目约定逻辑删除: del_flag='0' 为正常数据(COALESCE 兼容历史 NULL 数据)
        where += " AND COALESCE(del_flag, '0') = '0'"
    return where, params


def convert_row(row):
    """数据库行 -> Excel行

    - 日期统一转 YYYY-MM-DD 字符串(避免时区/序列号歧义);
    - 字符串剔除 xlsx 规范禁止的控制字符(\x00-\x08/\x0b/\x0c/\x0e-\x1f),
      生产同步数据中可能夹杂此类脏字符, 不剔除会引发 IllegalCharacterError;
    - None 保持空单元格。
    """
    clean = []
    for v in row:
        if isinstance(v, date):
            v = v.strftime('%Y-%m-%d')
        elif isinstance(v, str):
            v = ILLEGAL_CHARACTERS_RE.sub('', v)
        clean.append(v)
    return clean


def fmt_secs(seconds):
    """秒数 -> 中文可读时长"""
    seconds = int(seconds)
    h, rem = divmod(seconds, 3600)
    m, s = divmod(rem, 60)
    if h:
        return '%d时%02d分%02d秒' % (h, m, s)
    if m:
        return '%d分%02d秒' % (m, s)
    return '%d秒' % s


def load_checkpoint(path):
    """读取断点记录(损坏/缺失时返回 None, 不影响从头导出); 兼容 v1(last_id) 与 v2(ids_pos)"""
    try:
        with open(path, 'r', encoding='utf-8') as f:
            state = json.load(f)
        if 'rows_done' in state and 'file_idx' in state and ('ids_pos' in state or 'last_id' in state):
            return state
        LOG.warning('断点文件内容不完整, 忽略: %s', path)
    except (OSError, ValueError):
        LOG.warning('读取断点文件失败, 忽略: %s', path)
    return None


def save_checkpoint(path, **fields):
    """原子写入断点记录(临时文件 + os.replace, 避免断电等留下半成品 json)"""
    data = dict(fields)
    data['updated_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    tmp = path + '.tmp'
    with open(tmp, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    os.replace(tmp, path)


def collect_target_ids(session, where_sql, params):
    """一次性收集全部目标行 id: 单条语句顺序扫描(仅 int 列, 数秒~几十秒完成),
    autocommit 下为语句级短事务, 无长事务; 客户端内存约百MB以内"""
    sql = 'SELECT id FROM %s %s' % (TABLE_NAME, where_sql)
    with session['conn'].cursor() as cur:
        cur.execute(sql, params)
        ids = sorted(r[0] for r in cur.fetchall())
    return ids


def load_or_collect_ids(session, ids_file, where_sql, params):
    """优先复用已持久化的 id 清单(中断重跑免重复扫描), 否则收集并原子落盘"""
    if os.path.exists(ids_file):
        with open(ids_file, 'r', encoding='utf-8') as f:
            ids = [int(line) for line in f if line.strip()]
        LOG.info('复用已有 id 清单: %s (%d 个)', ids_file, len(ids))
        return ids
    LOG.info('阶段1/2: 顺序扫描收集全部目标行 id ...')
    ids = collect_target_ids(session, where_sql, params)
    tmp = ids_file + '.tmp'
    with open(tmp, 'w', encoding='utf-8') as f:
        f.write('\n'.join(str(i) for i in ids))
    os.replace(tmp, ids_file)
    LOG.info('id 清单已保存: %s (%d 个)', ids_file, len(ids))
    return ids


def fetch_rows_by_ids_with_retry(session, sql, id_chunk):
    """按主键 id 点查一批完整行数据(无稀疏区无效扫描), 连接中断自动重连重试;
    点查幂等, 重试安全"""
    delay = 10
    attempt = 0
    while True:
        try:
            with session['conn'].cursor() as cur:
                cur.execute(sql, (id_chunk,))
                return cur.fetchall()
        except psycopg2.OperationalError as exc:
            attempt += 1
            if attempt > session['max_retries']:
                raise
            LOG.warning('数据库连接中断(%s), %d 秒后重连重试(第 %d/%d 次)',
                        exc, delay, attempt, session['max_retries'])
            time.sleep(delay)
            delay *= 3
            try:
                session['conn'].close()
            except Exception:
                pass
            session['conn'] = psycopg2.connect(**session['kwargs'])
            session['conn'].autocommit = True


def do_export(args):
    db_cfg = DB_CONFIG  # 严格使用内置硬编码连接配置, 不接受任何外部覆盖
    os.makedirs(args.output, exist_ok=True)

    where_sql, params = build_where(args.start, args.end, args.include_deleted)
    # 两阶段导出:
    #   阶段1 一条短事务顺序扫描收集全部目标 id(数秒级, 仅 int 列);
    #   阶段2 按主键 id = ANY(清单切片) 精确点查整行, 无稀疏区无效扫描,
    #        每批独立短事务, 断点 = 清单偏移量(ids_pos), 天然支持续传
    select_sql = 'SELECT %s FROM %s WHERE id = ANY(%%s) ORDER BY id' % (
        ', '.join(c for c, _, _ in FIELDS), TABLE_NAME)

    conn_kwargs = dict(
        host=db_cfg['host'], port=db_cfg['port'], user=db_cfg['user'],
        password=db_cfg['password'], dbname=db_cfg['dbname'],
        connect_timeout=10, application_name='xc_export_bom_info')
    LOG.info('连接数据库 %s:%s/%s ...', db_cfg['host'], db_cfg['port'], db_cfg['dbname'])
    conn = psycopg2.connect(**conn_kwargs)
    conn.autocommit = True  # 每批查询独立短事务(只读), 不阻碍 autovacuum 回收
    session = {'conn': conn, 'kwargs': conn_kwargs, 'max_retries': args.max_retries}

    try:
        # 断点恢复: 上次已封卷的文件保留, 从其末行 id 之后继续
        ckpt_path = os.path.join(args.output, 'export_checkpoint.json')
        state = None
        if os.path.exists(ckpt_path):
            if args.restart:
                os.remove(ckpt_path)
                LOG.info('已按 --restart 删除断点记录, 从头导出')
            else:
                state = load_checkpoint(ckpt_path)
                if state:
                    LOG.info('检测到断点记录: 已完成 %d 行(分卷至第 %d 卷), 继续导出',
                             state['rows_done'], state['file_idx'])

        # 阶段1: id 清单(优先复用已落盘清单, 免重复全表扫描)
        ids_file = os.path.join(args.output, 'target_ids_%s_%s.txt' % (
            args.start.strftime('%Y%m%d'), args.end.strftime('%Y%m%d')))
        ids = load_or_collect_ids(session, ids_file, where_sql, params)
        total = len(ids)
        LOG.info('目标行总数: %d', total)
        if total == 0 and not state:
            LOG.warning('无符合条件的数据, 不生成文件')
            return 0

        # 断点偏移: v2 直接用 ids_pos; v1(last_id) 二分定位自动转换
        if state:
            if 'ids_pos' in state:
                pos = state['ids_pos']
            else:
                pos = bisect.bisect_right(ids, state['last_id'])
                save_checkpoint(ckpt_path, ids_pos=pos, rows_done=state['rows_done'],
                                file_idx=state['file_idx'])
                LOG.info('旧断点已转换: last_id=%d -> 清单偏移 %d', state['last_id'], pos)
            rows_done = state['rows_done']
        else:
            pos, rows_done = 0, 0
        session_start = rows_done  # 本次会话起点(速率按本次实际导出统计)

        base_name = 'sn_service_bom_info_%s-%s' % (args.start.strftime('%Y%m%d'), args.end.strftime('%Y%m%d'))
        writer = RollingExcelWriter(
            output_dir=args.output, base_name=base_name,
            headers=[h for _, h, _ in FIELDS], widths=[w for _, _, w in FIELDS],
            rows_per_file=args.rows_per_file,
            start_file_idx=state['file_idx'] if state else 1)

        t0 = time.time()
        files_done = len(writer.files)
        batch_no = 0
        t_batch0 = time.time()
        try:
            while pos < total:
                chunk = ids[pos:pos + args.batch_size]
                rows = fetch_rows_by_ids_with_retry(session, select_sql, chunk)
                batch_no += 1
                fetch_took = time.time() - t_batch0
                pos += len(chunk)
                if rows:
                    writer.write_rows(convert_row(r) for r in rows)
                    rows_done += len(rows)
                # 每批心跳: "拉取耗时"用于区分数据库慢与本地写入慢
                session_done = rows_done - session_start
                elapsed = max(time.time() - t0, 1e-6)
                speed = session_done / elapsed
                eta = (total - pos) / speed if speed > 0 else 0
                LOG.info('第 %d 批: %d 行 | 拉取耗时 %s | 清单 %d/%d (%.1f%%) | 速率 %.0f 行/秒 | 剩余约 %s',
                         batch_no, len(rows), fmt_secs(fetch_took), pos, total,
                         pos * 100.0 / total, speed, fmt_secs(eta))
                t_batch0 = time.time()
                if len(writer.files) > files_done:
                    # 发生封卷: 立即持久化断点(卷边界可靠; 未封卷部分中断后需重导)
                    files_done = len(writer.files)
                    save_checkpoint(ckpt_path, ids_pos=pos, rows_done=rows_done,
                                    file_idx=writer._file_idx)
            writer.close()
        except BaseException:
            # 任何异常(含用户中断)都不遗留半成品文件; 已封卷文件与断点记录保留供续传
            writer.abort()
            raise

        # 全部导出成功, 清除断点记录与 id 清单
        if os.path.exists(ckpt_path):
            os.remove(ckpt_path)
        if os.path.exists(ids_file):
            os.remove(ids_file)

        elapsed = time.time() - t0
        LOG.info('=' * 64)
        LOG.info('导出完成: 本次导出 %d 行, 累计 %d 行, 本次生成 %d 个文件, 总耗时 %s, 平均 %.0f 行/秒',
                 rows_done - session_start, rows_done, len(writer.files), fmt_secs(elapsed),
                 (rows_done - session_start) / elapsed if elapsed > 0 else 0)
        for path, n in writer.files:
            LOG.info('  %s <- %d 行', path, n)
        LOG.info('筛选区间: %s ~ %s | 数据库: %s', args.start, args.end, db_cfg['dbname'])
        return 0
    finally:
        try:
            session['conn'].close()
        except Exception:
            pass


def main():
    logging.basicConfig(level=logging.INFO,
                        format='%(asctime)s [%(levelname)s] %(message)s',
                        datefmt='%Y-%m-%d %H:%M:%S')
    args = parse_args()
    try:
        return do_export(args)
    except KeyboardInterrupt:
        LOG.error('用户中断导出, 未完成的临时文件已清理')
        return 130
    except psycopg2.OperationalError:
        LOG.exception('数据库连接中断且重试已耗尽, 断点已保留, 重新运行脚本可从断点继续导出')
        return 2
    except psycopg2.Error:
        LOG.exception('数据库查询/读取异常')
        return 3
    except OSError:
        LOG.exception('文件写入异常(请检查磁盘空间/目录权限)')
        return 4


if __name__ == '__main__':
    sys.exit(main())

